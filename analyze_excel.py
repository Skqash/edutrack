import openpyxl
import json

# Load the Excel file
xlsx_path = r"c:\laragon\www\edutrack\grading system\grading system excel.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
print('Sheet names:', wb.sheetnames)
print()

# Analyze each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'Sheet: {sheet_name}')
    print(f'Dimensions: {ws.dimensions}')
    print()
    
    # Print first 20 rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 20:
            print(row)
        else:
            break
    print('---')
    print()
