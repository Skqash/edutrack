import openpyxl
from openpyxl.utils import get_column_letter

# Load the workbook
wb = openpyxl.load_workbook('grading system/grading system excel.xlsx')
ws = wb.active

# Get basic info
print(f'Sheet Name: {ws.title}')
print(f'Dimensions: {ws.dimensions}')
print(f'Max Row: {ws.max_row}, Max Column: {ws.max_column}')
print()

# Get headers
print('=== HEADERS ===')
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)
print(f'Headers: {headers}')
print(f'Total columns: {len(headers)}')
print()

# Get all data
print('=== ALL DATA ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
    if i == 1:
        print(f'Row {i} (Header): {row}')
    else:
        print(f'Row {i}: {row}')

print()
print('=== ANALYSIS COMPLETE ===')
